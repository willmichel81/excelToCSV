# what this script will try and accomplish:
# import excel file
# read certain cells in file
# create csv based off this information
import os
from openpyxl import load_workbook
import json
import csv

class excelToCSV():

    def __init__(self,filepath):
        # user specified filepath with filename
        self.filepath = filepath
        # sees if file is there
        self.file_exists = self.isFileThere()
        # get file info (name,extension)
        self.file_info = self.fileInfo()

    # gets path and extension from file
    def fileInfo(self,filepath=""):
        if filepath == "":
            # if no file name is passed directly to "fileInfo()" use the file name that is passed when initializing class
            filepath = self.filepath
            # splits path to file[0] = path and file[1] = extension
        file_info = os.path.splitext(filepath)
        return file_info

    #Checks to see if file is there.
    def isFileThere(self, filepath = ""):
        if filepath == "":
            # if no file name is passed directly to "isFileThere()" use the file name that is passed when initializing class
            filepath = self.filepath
        file_exists = os.path.isfile(filepath);
        return file_exists

    def isDirThere(self, dir ="/"):
        dir_exists = os.path.isdir(dir);
        return dir_exists
    #excelOBJ: is the load_workbook funciton from openpyxl with the xlsx file passed to it.
    # sheet: that you wish to access
    # range: read from point a to point b
    # NOTE: if looking for a single sale pass in the format of "B2:B2"
    def readSheet(self, excelOBJ, sheet, range):
        print("Reading '{}' sheet from {} from Excel document ".format(sheet,range))
        # placeholder for dataset
        dataset = []
        # access excelObJ list with index of selected sheet
        sheet_ranges = excelOBJ[sheet]
        # loop through cells in given range "A4:A118..ect"
        for cells in sheet_ranges[range]:
            for cell in cells:
                # append cell value to empty dataset
                dataset.append(cell.value)
        return dataset

    def workBook(self, filepath=""):
        print("Creating workbook")
        if filepath == "":
            filepath = self.filepath
        if self.file_exists == True:
            wb = load_workbook(self.filepath)
        return wb
    # This is where the magic happens ... for the mostpart
    #create a json object based on data that was read from excel
    def createJSON(self,length = 0,datasetName = "record",datasetDataName = [],data = "",filepath = ""):
            i = 0
            x = 0
            print("Creating Json")
            json_str = '{'

            # loop through each elment and create  a json string foreach element
            while i < int(length):
                json_str += '"{}{}":'.format(datasetName,i)
                json_str += '{'
                for element in datasetDataName:
                    if x == len(data) - 1:
                        json_str += '"{}":'.format(element)
                        json_str += '"{}"'.format(data[x][i])
                        x = 0
                    else:
                        json_str += '"{}":'.format(element)
                        json_str += '"{}"'.format(data[x][i])
                        json_str += ','
                        x += 1
                # checks to see if it is the last elment. If not add "},", if yes add "}"
                if i == int(length) - 1:
                    json_str += '}'
                else:
                    json_str += '},'


                i += 1
            # end of while loop there for end of json
            json_str += '}'
            # parse json and save into a dict
            json_dict = json.loads(json_str)
            print("Finished creating Json")
            return json_dict

    # creates the csv off of the data found in excel
    def createCSV(self,fieldnames,dict="",dir="/",filename="import.csv"):
        print("Checking for directory")
        if(self.isDirThere(dir) == True):
            print("Found Dir")
            path = dir + filename
            # opens file with write access
            with open(path,mode='w') as csv_file:
                print("Creating {}".format(path))
                # writes the first line ... the header fieldnames
                writer = csv.DictWriter(csv_file, fieldnames = fieldnames)
                writer.writeheader()
                # writes the data for each
                for key,value in dict.items():
                     writer.writerow(value)
            print("Finished writing to: \n{}".format(path))
        else:
            print("{} was not found... Creating directory".format(dir))
            os.mkdir(dir)
            print("Directory created")
            self.createCSV(fieldnames,dict,dir,filename)
